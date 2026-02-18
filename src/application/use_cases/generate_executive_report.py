
import duckdb
import io
import datetime
import calendar
from typing import Dict, Any, List
import pandas as pd
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as PlatypusImage, PageBreak, Frame, PageTemplate
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.lib.units import inch
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import openpyxl
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from .generate_origin_report import GenerateOriginReport
from .generate_destination_report import GenerateDestinationReport
from .generate_time_report import GenerateTimeReport
from .generate_flight_type_report import GenerateFlightTypeReport
from .generate_company_report import GenerateCompanyReport
from .generate_region_report import GenerateRegionReport
from .generate_heatmap_report import GenerateHeatmapReport

class GenerateExecutiveReport:
    def __init__(self, db_path: str = "data/metrics.duckdb"):
        self.db_path = db_path
        self.origin_uc = GenerateOriginReport(db_path)
        self.dest_uc = GenerateDestinationReport(db_path)
        self.time_uc = GenerateTimeReport(db_path)
        self.type_uc = GenerateFlightTypeReport(db_path)
        self.company_uc = GenerateCompanyReport(db_path)
        self.region_uc = GenerateRegionReport(db_path)
        self.heatmap_uc = GenerateHeatmapReport(db_path)

    def _get_filter_summary(self, filters: Dict[str, Any]) -> List[List[str]]:
        summary = []
        
        # Helper to format list items
        def format_list(items):
            if not items: return "Todos"
            # Extract labels if items are objects, or use values
            values = []
            for i in items:
                if isinstance(i, dict):
                    if 'label' in i: values.append(i['label'])
                    elif 'value' in i:
                         v = i['value']
                         if isinstance(v, dict) and 'icao_code' in v: values.append(v['icao_code'])
                         else: values.append(str(v))
                    else: values.append(str(i))
                else: values.append(str(i))
            return ", ".join(values)

        # Date Range
        start = filters.get('start_date')
        end = filters.get('end_date')
        date_str = "Todo el periodo disponible"
        if start and end: date_str = f"{start} al {end}"
        elif start: date_str = f"Desde {start}"
        elif end: date_str = f"Hasta {end}"
        summary.append(["Rango de Fechas", date_str])

        # Levels
        min_lvl = filters.get('min_level')
        max_lvl = filters.get('max_level')
        lvl_str = "Todos"
        if min_lvl and max_lvl: lvl_str = f"{min_lvl} - {max_lvl}"
        elif min_lvl: lvl_str = f"Min: {min_lvl}"
        elif max_lvl: lvl_str = f"Max: {max_lvl}"
        summary.append(["Niveles de Vuelo", lvl_str])

        # Lists
        summary.append(["Orígenes", format_list(filters.get('selectedOrigins'))])
        summary.append(["Destinos", format_list(filters.get('selectedDestinations'))])
        summary.append(["Aerolíneas", format_list(filters.get('selectedEmpresa'))])
        summary.append(["Tipos de Aeronave", format_list(filters.get('selectedTipoAeronave'))])
        summary.append(["Tipos de Vuelo", format_list(filters.get('selectedTipoVuelo'))])
        summary.append(["Callsigns", format_list(filters.get('selectedCallsign'))])
        summary.append(["Matrículas", format_list(filters.get('selectedMetricula'))])

        return summary

    def _get_aggregated_data(self, filters: Dict[str, Any]) -> Dict[str, Any]:
        data = {}
        try:
            # 1. Basic Distributions
            data['origins'] = self.origin_uc._get_data(filters)
            data['destinations'] = self.dest_uc._get_data(filters)
            
            # 2. Time Evolution (Month and Year)
            # data['time_month'] is standard format [{'period': '2023-01', 'value': 100}, ...]
            data['time_month'] = self.time_uc._get_data(filters, 'month')
            
            # For Year, we reuse time_uc logic but grouped by year if possible, 
            # or manual aggregation if _get_data only supports 'month'.
            # Let's assume for now we aggregate localy from month data if year not supported directly,
            # BUT efficient SQL is better. Let's try to pass 'year' if supported, else aggregate.
            # Looking at GenerateTimeReport (not visible now, but standard implementation), let's aggregate here just in case.
            df_time = pd.DataFrame(data['time_month'])
            if not df_time.empty:
                # Fix: GenerateTimeReport returns 'name' as key for period, not 'period'
                df_time['year'] = df_time['name'].astype(str).str[:4]
                data['time_year'] = df_time.groupby('year')['value'].sum().reset_index().rename(columns={'year':'period'}).to_dict('records')
            else:
                data['time_year'] = []

            # 3. Categorical
            data['types'] = self.type_uc._get_data(filters)
            data['companies'] = self.company_uc._get_data(filters)
            
            # 4. Regional
            data['regions_origin'] = self.region_uc._get_data(filters, 'origin')
            data['regions_dest'] = self.region_uc._get_data(filters, 'destination')
            
            # 5. Heatmaps
            data['heatmap_dep'] = self.heatmap_uc._get_data(filters, 'hora_salida')
            data['heatmap_arr'] = self.heatmap_uc._get_data(filters, 'hora_llegada')

            # 6. KPIs
            data['total_flights'] = sum(d['value'] for d in data['time_month'])
            
            if data['origins']:
                top_ori = max(data['origins'], key=lambda x: x['count'])
                data['top_origin'] = f"{top_ori['origen']} ({top_ori['count']})"
            else: data['top_origin'] = "N/A"
            
            if data['destinations']:
                top_des = max(data['destinations'], key=lambda x: x['count'])
                data['top_dest'] = f"{top_des['destino']} ({top_des['count']})"
            else: data['top_dest'] = "N/A"
            
            # Top Company
            if data['companies']:
                top_comp = max(data['companies'], key=lambda x: x['value'])
                data['top_company'] = f"{top_comp['name']} ({top_comp['value']})"
            else: data['top_company'] = "N/A"

            return data
        except Exception as e:
            print(f"Error gathering data: {e}")
            raise e

    def generate_pdf(self, filters: Dict[str, Any]) -> io.BytesIO:
        data = self._get_aggregated_data(filters)
        buffer = io.BytesIO()
        
        # Document Setup
        doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom Styles
        title_style = ParagraphStyle('ReportTitle', parent=styles['Title'], fontSize=26, textColor=colors.HexColor('#111827'), spaceAfter=10, fontName='Helvetica-Bold')
        subtitle_style = ParagraphStyle('ReportSubtitle', parent=styles['Heading2'], fontSize=16, textColor=colors.HexColor('#374151'), spaceBefore=20, spaceAfter=10, fontName='Helvetica-Bold', borderPadding=5, borderColor=colors.HexColor('#e5e7eb'), borderWidth=0, borderBottomWidth=1)
        text_style = ParagraphStyle('ReportText', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#4b5563'))
        
        # --- HEADER ---
        elements.append(Paragraph("Informe Ejecutivo de Operaciones Aéreas", title_style))
        elements.append(Paragraph(f"Fecha de Generación: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}", text_style))
        elements.append(Spacer(1, 20))

        # --- FILTERS SUMMARY ---
        elements.append(Paragraph("Filtros Aplicados", subtitle_style))
        filter_data = self._get_filter_summary(filters)
        t_filters = Table(filter_data, colWidths=[150, 350])
        t_filters.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (0,-1), colors.HexColor('#f3f4f6')),
            ('TEXTCOLOR', (0,0), (0,-1), colors.HexColor('#1f2937')),
            ('ACellPadding', (0,0), (-1,-1), 6),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e5e7eb')),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 9),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        elements.append(t_filters)
        elements.append(Spacer(1, 20))

        # --- KEY METRICS (KPIs) ---
        elements.append(Paragraph("Métricas Clave", subtitle_style))
        kpi_data = [
            ["TOTAL VUELOS", "TOP ORIGEN", "TOP DESTINO", "TOP AEROLÍNEA"],
            [f"{data['total_flights']:,}", data['top_origin'], data['top_dest'], data['top_company']]
        ]
        t_kpis = Table(kpi_data, colWidths=[130, 130, 130, 130])
        t_kpis.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#4f46e5')), # Indigo 600
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 10),
            ('FONTSIZE', (0,1), (-1,1), 12),
            ('BOTTOMPADDING', (0,0), (-1,-1), 12),
            ('TOPPADDING', (0,0), (-1,-1), 12),
            ('BOX', (0,0), (-1,-1), 1, colors.HexColor('#4f46e5')),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#e0e7ff')),
            ('BACKGROUND', (0,1), (-1,1), colors.HexColor('#eef2ff')), # Indigo 50
        ]))
        elements.append(t_kpis)
        elements.append(Spacer(1, 20))

        # --- PAGE 1: TEMPORAL EVOLUTION ---
        elements.append(Paragraph("Evolución Temporal", subtitle_style))
        
        # Monthly Chart
        img_month = self.time_uc._generate_chart(data['time_month'], 'month')
        if img_month:
            elements.append(PlatypusImage(img_month, width=500, height=250))
            elements.append(Paragraph("Tendencia mensual de vuelos.", text_style))
            elements.append(Spacer(1, 10))
            
        # Yearly Chart (Generated manually as reuse might fail if not designed for it)
        if data['time_year']:
            # Create yearly chart on the fly
            try:
                plt.figure(figsize=(10, 4))
                df_y = pd.DataFrame(data['time_year'])
                plt.bar(df_y['period'], df_y['value'], color='#3b82f6')
                plt.title("Vuelos por Año")
                plt.grid(axis='y', linestyle='--', alpha=0.5)
                buf_y = io.BytesIO()
                plt.savefig(buf_y, format='png', bbox_inches='tight', dpi=100)
                plt.close()
                buf_y.seek(0)
                elements.append(PlatypusImage(buf_y, width=500, height=200))
            except Exception as e:
                print(f"Error gen yearly chart: {e}")

        elements.append(PageBreak())

        # --- PAGE 2: GEOGRAPHIC DISTRIBUTION ---
        elements.append(Paragraph("Distribución Geográfica", subtitle_style))
        
        # Origins and Destinations Tables
        t_orig_data = [['Top Orígenes', 'Vuelos']] + [[x['origen'], f"{x['count']:,}"] for x in sorted(data['origins'], key=lambda x: x['count'], reverse=True)[:10]]
        t_dest_data = [['Top Destinos', 'Vuelos']] + [[x['destino'], f"{x['count']:,}"] for x in sorted(data['destinations'], key=lambda x: x['count'], reverse=True)[:10]]
        
        t_orig = Table(t_orig_data, colWidths=[150, 80])
        t_dest = Table(t_dest_data, colWidths=[150, 80])
        
        grid_style = TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#374151')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#d1d5db')),
            ('ALIGN', (1,0), (1,-1), 'RIGHT'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f9fafb')]),
        ])
        t_orig.setStyle(grid_style)
        t_dest.setStyle(grid_style)
        
        elements.append(Table([[t_orig, Spacer(20,0), t_dest]], hAlign='CENTER'))
        elements.append(Spacer(1, 20))

        # Regional Charts
        elements.append(Paragraph("Análisis Regional (Treemaps)", subtitle_style))
        
        img_reg_ori = self.region_uc._generate_chart(data['regions_origin'])
        img_reg_dest = self.region_uc._generate_chart(data['regions_dest'])
        
        row_reg = []
        if img_reg_ori: row_reg.append(PlatypusImage(img_reg_ori, width=250, height=200))
        if img_reg_dest: row_reg.append(PlatypusImage(img_reg_dest, width=250, height=200))
        
        if row_reg:
            elements.append(Table([row_reg], hAlign='CENTER'))
            elements.append(Paragraph("Izquierda: Región Origen | Derecha: Región Destino", ParagraphStyle('Caption', parent=styles['Normal'], alignment=TA_CENTER, fontSize=9, textColor=colors.gray)))

        elements.append(PageBreak())

        # --- PAGE 3: OPERATIONAL ANALYSIS ---
        elements.append(Paragraph("Análisis Operativo", subtitle_style))
        
        # Flight Types
        img_type = self.type_uc._generate_chart(data['types'])
        if img_type:
            elements.append(Paragraph("Tipos de Vuelo", styles['Heading3']))
            elements.append(PlatypusImage(img_type, width=400, height=250))
            elements.append(Spacer(1, 10))

        # Companies
        img_comp = self.company_uc._generate_chart(data['companies'][:15]) # Slice data!
        if img_comp:
            elements.append(Paragraph("Top 15 Aerolíneas", styles['Heading3']))
            elements.append(PlatypusImage(img_comp, width=500, height=300))
            
        elements.append(PageBreak())
        
        # --- PAGE 4: HEATMAPS ---
        elements.append(Paragraph("Mapas de Calor (Frecuencia Semanal/Horaria)", subtitle_style))
        
        img_heat_dep = self.heatmap_uc._generate_chart(data['heatmap_dep'])
        if img_heat_dep:
            elements.append(Paragraph("Hora de Salida", styles['Heading3']))
            elements.append(PlatypusImage(img_heat_dep, width=500, height=250))
            elements.append(Spacer(1, 15))
            
        img_heat_arr = self.heatmap_uc._generate_chart(data['heatmap_arr'])
        if img_heat_arr:
            elements.append(Paragraph("Hora de Llegada", styles['Heading3']))
            elements.append(PlatypusImage(img_heat_arr, width=500, height=250))

        doc.build(elements)
        buffer.seek(0)
        return buffer

    def generate_excel(self, filters: Dict[str, Any]) -> io.BytesIO:
        data = self._get_aggregated_data(filters)
        output = io.BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        workbook = writer.book
        
        # Styles
        title_font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1e3a8a', end_color='1e3a8a', fill_type='solid') # Dark Blue
        kpi_font = Font(name='Arial', size=12, bold=True)
        
        # --- DASHBOARD SHEET ---
        ws = workbook.create_sheet("Executive Dashboard")
        workbook.active = ws
        ws.sheet_view.showGridLines = False
        
        # Header Banner
        ws.merge_cells('A1:L2')
        cell = ws['A1']
        cell.value = "INFORME EJECUTIVO DE OPERACIONES AÉREAS"
        cell.font = title_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws['A3'] = f"Generado: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}"
        
        # KPIs Row
        kpis = [
            ("Total Vuelos", data['total_flights']),
            ("Top Origen", data['top_origin']),
            ("Top Destino", data['top_dest']),
            ("Top Aerolínea", data['top_company'])
        ]
        
        start_col = 2
        for title, value in kpis:
            ws.cell(row=5, column=start_col).value = title
            ws.cell(row=5, column=start_col).font = Font(bold=True, color='6b7280')
            ws.cell(row=6, column=start_col).value = value
            ws.cell(row=6, column=start_col).font = Font(size=14, bold=True, color='1e3a8a')
            start_col += 3
            
        # Filters Summary (Side Panel)
        ws.merge_cells('A8:B8')
        ws['A8'] = "Filtros Aplicados"
        ws['A8'].font = Font(bold=True, underline='single')
        
        row = 9
        for label, val in self._get_filter_summary(filters):
            ws.cell(row=row, column=1).value = label
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=2).value = val
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
            row += 1
            
        # Charts Positioning
        # Time Chart
        img_time = self.time_uc._generate_chart(data['time_month'], 'month')
        if img_time:
            ws.add_image(ExcelImage(img_time), 'D9')
            
        # Type Chart
        img_type = self.type_uc._generate_chart(data['types'])
        if img_type:
            ws.add_image(ExcelImage(img_type), 'D25')
            
        # Company Chart
        img_comp = self.company_uc._generate_chart(data['companies'][:10])
        if img_comp:
            ws.add_image(ExcelImage(img_comp), 'H25')
            
        # --- DATA SHEETS ---
        def write_sheet(name, data_dict):
            if not data_dict: return
            df = pd.DataFrame(data_dict)
            df.to_excel(writer, sheet_name=name, index=False)
            # Auto-adjust columns
            worksheet = writer.sheets[name]
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except: pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        write_sheet("Orígenes", data['origins'])
        write_sheet("Destinos", data['destinations'])
        write_sheet("Tiempo", data['time_month'])
        write_sheet("Empresas", data['companies'])
        write_sheet("Regiones Origen", data['regions_origin'])
        write_sheet("Regiones Destino", data['regions_dest'])
        write_sheet("Mapa Calor Salida", data['heatmap_dep'])
        write_sheet("Mapa Calor Llegada", data['heatmap_arr'])

        writer.close()
        output.seek(0)
        return output
